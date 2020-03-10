# 银行登录系统小程序
# 先输入用户名，若不正确将重新登陆，直到正确为止。
# 再输入密码，只有3次输入机会，都错误跳出系统。
# 登陆成功后输入取款金额，若金额不足将选择是否继续还是退出。
# 选择是，重新输入取款金额；如果退出，调出系统。
# 如果取款成功将返回剩余帐号金额，并且退出系统。
# 取款金额是100的整数倍。
# dic = {
# 'tom':[123456, 1000],
# 'joe':[246810, 200]
# }

import time
import os
import openpyxl

class User(object): # 定义用户类，可以完成：2、登陆；3、取款；4、存款；5、注销
    user_dic = {}   # 创建用户字典，账户名：对象地址 一一对应
    def __init__(self, account, password, balance = 0, status = 0, identity = 0, frozen = 0):    # 初始化信息，有用户，密码，余额，登陆状态
        self.account = account
        self.password = password
        self.balance = balance
        self.status = status        # status = 0 为未登录；status = 1 为登陆。
        self.identity = identity    # 0为普通，1位贵宾
        self.frozen = frozen        # 账户状态，0为正常，1为冻结
        self.user_dic[account] = self       # 记录对象地址

    def log_in(self):   # 登陆
        count = 0
        while count < 3:
            try:
                # input_account = input('请输入账户名：')
                input_password = input('请输入密码：')
            except:
                pass
            if input_password == 'exit':
                print('再见！')
                exit()
            if input_password == self.password:
                print('确认中，请稍候...')         # 装逼专用
                time.sleep(1)
                if self.frozen == 1:
                    print('对不起，您的账户已被锁定。\n请致电13816917340获取详细信息。\n--------------------------')
                    exit()
                else:
                    pass
                # print(1)            # 调试检测专用
                if self.identity == 0:      # 普通用户登陆
                    print('登陆成功，欢迎您，%s！\n--------------------------' %self.account)
                else:                       # 贵宾登陆
                    print('登陆成功，欢迎您，尊敬的贵宾%s！\n--------------------------' %self.account)
                self.status = 1
                time.sleep(1)
                break
            else:
                count += 1
                print('密码错误，请重试！输入“exit”退出\n--------------------------')
            if count == 3:
                print('登陆失败3次，界面锁定！')
                exit()

    def draw(self):     # 取款
        if self.status != 1:        # 判断登录状态
            print('您尚未登录，请登陆！\n--------------------------')
            self.log_in()
        while True:     # 取款循环
            try:
                print('您的余额为%s' %self.balance)
                cash = int(input('请输入取款数额，必须为100的整数倍，输入“0”退出：'))
                if cash == 0:
                    i = os.system('cls')
                    break
                print('确认中，请稍候...')
                time.sleep(2)
                i = os.system('cls')
                if cash % 100 != 0:
                    print('金额输入错误，请重试\n--------------------------')
                    continue
                elif cash > self.balance:
                    print('余额不足，请重试\n--------------------------')
                    continue
                else:
                    self.balance = self.balance - cash
                    print('取款成功！您的余额为%s\n--------------------------' %self.balance)
                    break
            except:
                print('金额输入错误，请重试\n--------------------------')
                continue
        return self.balance     # 返回余额


    def top_up(self):   # 存款
        if self.status != 1:        # 判断登录状态
            print('您尚未登录，请登陆！\n--------------------------')
            self.log_in()
        while True:
            try:
                print('您的余额为%s' %self.balance)
                money = int(input('请输入存款数，必须为100的整数倍，输入“0”退出：'))
                # print(1)        # 调试检验专用
                if money == 0:
                    # print(2)        # 调试检验专用
                    i = os.system('cls')
                    break       # 输入0跳出循环
                print('确认中，请稍后...')
                time.sleep(2)
                if money % 100 != 0:
                    # print(3)        # 调试检验专用
                    print('金额输入错误，请重试\n--------------------------')
                    continue
                else:
                    i = os.system('cls')
                    self.balance += money
                    print('存款成功！您的余额为%s\n--------------------------' %self.balance)
                    break
                print('确认中，请稍后...')     # 装逼专用
                time.sleep(2)
            except:
                print('金额输入错误，请重试\n--------------------------')
                continue
        return (self.balance)

    def log_out(self):
        self.status = 0

class Manager(object):
    # 管理员分为1、2、3级
    # 1级权限: identity = 10, 仅可以查看所有用户及余额
    # 2级权限: identity = 20, 可以修改用户余额
    # 3级权限：identity = 30, 可以查看信息、修改余额，还可以冻结用户，添加、删除管理员
    manager_dic = {}
    def __init__(self, account, password, identity):    # 初始化数据，帐号，密码，身份证书
        self.account = account
        self.password = password
        self.identity = identity
        self.manager_dic[account] = self    # 储存身份信息



class Menu(object):         # 菜单栏，有5个功能
    # def __init__ (self):
    #     print('欢迎来到张子奇的神秘银行！')
    #     print('1、登陆账户')
    #     print('2、注册账户')
    #     print('3、取款')
    #     print('4、存款')
    #     print('5、注销')
    #     print('\n--------------------------')

    # 以下为从数据源获取信息的函数封装
    def get_account(self):                         # 从excel中获取用户列表
        wb = openpyxl.load_workbook('account_info.xlsx')
        sheet = wb['User']
        account_list = []
        rows = sheet.max_row
        for i in range(2, rows + 1):
            value = sheet.cell(row = i, column = 2).value
            account_list.append(value)
        return account_list

    def get_frozen(self, row):                      # 获取冻结状态
        wb = openpyxl.load_workbook('account_info.xlsx')
        sheet = wb['User']
        frozen = sheet.cell(row = row, column = 6).value
        return frozen

    def get_info(self, row):                         # 从excel中获取用户对应的信息
        wb = openpyxl.load_workbook('account_info.xlsx')
        sheet = wb['User']
        account_user = sheet.cell(row = row, column = 2).value
        password_user = str(sheet.cell(row = row, column = 3).value)
        balance_user = sheet.cell(row = row, column = 4).value
        identity_user = sheet.cell(row = row, column = 5).value
        frozen_user = sheet.cell(row = row, column = 6).value
        # print(account_user,password_user,balance_user,identity_user)    # 调试检验专用
        user = User(account = account_user, password = password_user,
                    balance = balance_user, identity = identity_user,
                    frozen = frozen_user)               # 生成 User对象

    def append_user(self, account, password, balance, identity):    # 注册成功后信息导入excel
        wb = openpyxl.load_workbook('account_info.xlsx')
        sheet = wb['User']
        rows = sheet.max_row            # 获取最大行数
        info = [rows, account, password, balance, identity]
        sheet.append(info)
        wb.save('account_info.xlsx')

    def update_balance(self, row, new_balance):            # 更新账户余额
        wb = openpyxl.load_workbook('account_info.xlsx')
        sheet = wb['User']
        sheet.cell(row = row, column = 4).value = new_balance
        wb.save('account_info.xlsx')

    # 这是个恶趣味问候语
    def greeting(self, name):               # 问候函数
        if User.user_dic[name].identity == 0:   # 0为普通会员
            print('欢迎您，%s！\n--------------------------' %name)
        else:                                   # 1为贵宾
            print('欢迎您，尊敬的贵宾%s！\n--------------------------' %name)

    # 以下为主菜单主函数
    def choice(self):       # 输入选择
        # print(account_list)     # 调试检验专用
        # account = 'Zzq'         # 调试检测转用
        while True:         # 死循环
            try:            # 问候语句，如果有account就问候，如果没有就没有
                self.greeting(account)
            except:
                pass
            try:
                print('欢迎来到张子奇的神秘银行！')
                print('1、登陆账户')
                print('2、注册账户')
                print('3、取款')
                print('4、存款')
                print('5、查询余额')
                print('6、注销')
                print('\n--------------------------')
                choose = input('请输入对应数字选择对应功能，输入“exit”退出：')
            except:
                print('输入有误，请重新输入！\n--------------------------')
                continue
            if choose == 'exit':    # 退出
                print('再见！')
                exit()

            elif choose == '1':     # 登陆
                i = os.system('cls')
                try:
                    if User.user_dic[account].status == 1:
                        print('对不起，%s已登录。请注销后再操作。' %account)
                except:
                    while True:
                        account_list = self.get_account()
                        # print(account_list)     # 调试检验专用
                        account = input('请输入账户：')
                        if account in account_list:         # 判断账户名是否存在
                            # 获取account信息并实例化
                            row = account_list.index(account) + 2
                            # print(row)                    # 调试检验专用
                            self.get_info(row)              # 找到对应信息并创建对象
                            # print(User.user_dic)          # 调试检验专用
                            User.user_dic[account].log_in() # 跳转到User.log_in()方法
                            i = os.system('cls')
                            break
                        else:
                            i = os.system('cls')
                            print('账户输入有误，请重试\n--------------------------')

            elif choose == '2':     # 注册
                i = os.system('cls')
                try:
                    if User.user_dic[account].status == 1:
                        print('对不起，%s已登录。请注销后再操作。' %account)
                except:
                    while True:         # 用户名输入循环
                        new_account = input('请输入用户名，输入“exit”退出：')
                        if new_account == 'exit':
                            break
                        if len(new_account) < 3:
                            print('用户名不得小于3位')
                            continue
                        elif new_account in User.user_dic:
                            print('用户名已存在，请重新输入！')
                            continue
                        else:
                            break       # 如无问题跳出循环
                    if new_account == 'exit':
                        i = os.system('cls')
                        continue
                    while True:         # 密码输入循环
                        new_password_1 = input('请输入密码，输入“exit”退出：')
                        if new_password_1 == 'exit':
                            break
                        if len(new_password_1) < 6:         # 判断：密码小于6位
                            print('密码不得小于6位，请重新输入。')
                            continue
                        new_password_2 = input('请确认密码：')
                        if new_password_2 != new_password_1:
                            print('密码输入有误，请重新输入。')
                        else:
                            break       # 如无问题跳出循环
                    if new_password_1 == 'exit':    # 退出注册环节
                        i = os.system('cls')
                        continue
                    # user = User(new_account, new_password_1, 0) #注册成功
                    self.append_user(new_account, new_password_1, balance = 0, identity = 0)
                    print('用户创建中，请稍候...')       #装逼专用
                    time.sleep(2)
                    input('恭喜您，用户创建成功！输入回车继续。')
                    i = os.system('cls')

            elif choose == '3':     # 取款
                i = os.system('cls')
                try:
                    balance = User.user_dic[account].draw()
                    row = account_list.index(account) + 2
                    self.update_balance(row, balance)
                except:
                    print('您尚未登录，请登陆！\n--------------------------')

            elif choose == '4':     # 存款
                i = os.system('cls')
                try:
                    # print(1)      # 调试检验专用
                    balance = User.user_dic[account].top_up()
                    # print(balance)  # 调试检验专用
                    row = account_list.index(account) + 2
                    self.update_balance(row, balance)
                except:
                    # print(2)      # 调试检验专用
                    print('您尚未登录，请登陆！\n--------------------------')

            elif choose == '5':     # 查询余额
                i = os.system('cls')
                try:                # 输入验证
                    print('您的余额为：%s元\n--------------------------' %User.user_dic[account].balance)
                except:             # 如果验证失败报错
                    print('您尚未登录，请登陆！\n--------------------------')

            elif choose == '6':     # 注销
                i = os.system('cls')
                try:
                    User.user_dic[account].log_out()        # 注销
                    account = None                          # 登录账户置0
                    User.user_dic = {}                      # 登陆用户字典置零，减少内存占用
                    # print(User.user_dic)                  # 调试检测专用

                except:
                    print('您尚未登录，请登陆！\n--------------------------')
            else:                   # 输入错误
                i = os.system('cls')
                print('输入有误，请重新输入！\n--------------------------')


def main():
    # user = User('Tom', '123456', 1000)                    # 调试检测专用
    # user = User('Joe', '246810', 200)                     # 调试检测专用
    # user = User('Zzq', '666666', 5000, identity = 1)      # 调试检测专用
    # user = User('Wyh', '777777', 3000, identity = 1)      # 调试检测专用
    # print(User.user_dic)                                  # 调试检测专用
    menu = Menu()
    menu.choice()


if __name__ == '__main__':
    main()
